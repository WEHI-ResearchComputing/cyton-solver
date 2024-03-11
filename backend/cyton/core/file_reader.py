"""
Last Edit: 11-Feb-2024

Extracts required information from the workbook.
"""
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from itertools import groupby
from cyton.core.utils import remove_empty
from typing import SupportsFloat, Any
from cyton.core.types import *
from datetime import datetime
from abc import ABC, abstractmethod

def get_worksheet(file: str) -> Worksheet:
	# Data_Only parameter required to remove equations
	workbook = openpyxl.load_workbook(file, data_only=True)
	sheet_names = workbook.sheetnames

	try:
		# Try Legacy data format by Cohort Explorer v2018.05.09
		ws = workbook['Input for CytonSolver']
	except Exception:
		# Else, default is first sheet of FormatHLSep2016
		ws = workbook[sheet_names[0]]  # Selects the first sheet
	return ws

class ReadData(ABC):
	condition_names: Conditions
	generation_per_condition: PerCond[MaxGeneration]
	harvested_times: PerCond[PerTime[HarvestTime]]
	harvested_times_reps: PerCond[Reps[HarvestTime]]
	num_time_points: NumTimePoints
	data: PerCond[PerTime[PerRep[PerGen[CellCount]]]]

	def __init__(self, file: str):
		ws = get_worksheet(file)

		# Extract meta information of the experiment
		self.meta_information = self.get_experiment_info(ws)
		self.condition_names = self.get_condition_names(ws)
		self.generation_per_condition = self.get_generation_information(ws)
		self.harvested_times, self.harvested_times_reps, self.num_time_points = self.get_time_points(ws)
		self.data = self.get_cell_number(ws)


	@staticmethod
	@abstractmethod
	def get_condition_names(worksheet: Worksheet) -> Conditions:
		...
	
	@abstractmethod
	def get_generation_information(self, worksheet: Worksheet) -> PerCond[MaxGeneration]:
		...

	@abstractmethod
	def get_time_points(self, worksheet: Worksheet) -> tuple[
		PerCond[PerTime[HarvestTime]], PerCond[Reps[HarvestTime]], NumTimePoints
	]:
		...

	@abstractmethod
	def get_cell_number(self, worksheet: Worksheet) -> PerCond[PerTime[PerRep[PerGen[CellCount]]]]:
		...

	@staticmethod
	@abstractmethod
	def get_experiment_info(worksheet: Worksheet) -> Any:
		...

def get_read_data(path: str) -> ReadData:
	worksheet = get_worksheet(path)

	format_tag = worksheet['B1']
	if format_tag.value == 'HL-SEP-2016':
		return ModernReadData(path)
	else:
		return LegacyReadData(path)

class LegacyReadData(ReadData):
	meta_information: LegacyExperimentMetadata

	@staticmethod
	def get_experiment_info(worksheet: Worksheet) -> LegacyExperimentMetadata:
		"""
		Collects user defined experiment information.

		:param1 worksheet: (Worksheet) openpyxl worksheet object
		:param2 data_format: (bool) True if we're using the new Excel data format
		:returns: (dict) experiments basic information
		:raises CustomError: raises a CustomError (inherits Exception) for missing critical information
		"""
		exp_info: ExperimentInfo | LegacyExperimentMetadata

		# Initialise information dictionary for legacy type data format
		exp_info = {
			'stimulus': '',
			'date': '',
			'cell_type': '',
			'comment': ''
		}
		# Iterate through first row [A1 - max_col1]
		col_idx = 0
		for row_cell_object in worksheet.iter_rows(min_col=1, max_col=worksheet.max_column, min_row=1, max_row=1):
			for cell in row_cell_object:
				header = cell.value
				if header is not None:
					value = row_cell_object[col_idx + 1].value
					if header == 'Stimulus' or header == 'stimulus':
						exp_info['stimulus'] = str(value)
					elif header == 'Date' or header == 'date':
						# If date is not specified default to current time
						if value is None:
							ts = str(datetime.now().replace(microsecond=0))
							exp_info['date'] = ts
						else:
							exp_info['date'] = str(value)
					elif header == 'Cell' or header == 'cell':
						exp_info['cell_type'] = str(value)
					elif header == 'Comment' or header == 'comment':
						exp_info['comment'] = str(value)
				else:
					if header is None:
						pass
					elif col_idx == 0:
						raise Exception("Missing stimulus name!")
					elif col_idx == 1:
						print("Missing experiment date time. Recording it to current time...")
						ts = str(datetime.now().replace(microsecond=0))
						exp_info['date'] = ts
					elif col_idx == 2:
						raise Exception("Missing cell type!")
					elif col_idx == 3:
						comment = 'No comment'
						exp_info['comment'] = comment
				col_idx += 1

		return exp_info

	@staticmethod
	def get_condition_names(worksheet: Worksheet) -> Conditions:
		"""
		Collects experiment condition names (e.g. Wildtype-1U, Wildtype-Unstim)

		:param1 worksheet: openpyxl worksheet object
		:param2 data_format: validation of data format
		:returns: condition names
		"""

		conditions = []
		for col_cell_object in worksheet.iter_cols(min_col=1, max_col=1, min_row=2, max_row=worksheet.max_row):
			for cell in col_cell_object:
				val = cell.value
				if val is not None:
					conditions.append(str(val))
		return conditions

	def get_generation_information(self, worksheet: Worksheet) -> PerCond[MaxGeneration]:
		"""
		Collects last generation information per conditions.

		:param1 worksheet: (Worksheet) openpyxl worksheet object
		:param2 data_format: (bool) validation of data format
		:returns: (list) last generation per condition indexed by exact order of condition list
		"""

		generation = []
		counter = 0
		for col_cell_object in worksheet.iter_cols(min_col=2, max_col=2, min_row=2, max_row=worksheet.max_row + 1):
			for cell in col_cell_object:
				val = cell.value
				if val is None:
					generation.append(counter - 2)
					counter = 0
				else:
					counter += 1
		return generation

	def get_time_points(self, worksheet: Worksheet) -> tuple[
		PerCond[PerTime[HarvestTime]], PerCond[Reps[HarvestTime]], NumTimePoints
	]:
		"""
		Collects harvested time points per conditions.

		:param1 worksheet: openpyxl worksheet object
		:param2 data_format: validation of data format
		:returns: (tuple; list, list) returns harvested time points & duplicates of harvested time points to represent replicates
		"""

		# Initialise with empty list with same dimension as get_condition_names. This allows to access time point data sorted by condition index.
		harvested_times_reps = [
			[] for _ in range(len(self.condition_names))
		]

		icnd = 0
		itpt = 2
		next_condition_break = self.generation_per_condition[icnd] + 3
		for row_cell_object in worksheet.iter_rows(min_col=2, max_col=worksheet.max_column, min_row=2, max_row=worksheet.max_row):
			for cell in row_cell_object:
				val = cell.value
				if (cell.row == itpt) and val is not None:
					harvested_times_reps[icnd].append(val)

			# A condition check for moving onto next condition & update icnd, itpt
			if row_cell_object[-1].row % next_condition_break == 0:
				itpt += self.generation_per_condition[icnd] + 3
				icnd += 1
				if icnd > len(self.condition_names) - 1:
					break
				next_condition_break += self.generation_per_condition[icnd] + 3

		harvested_times = [
			[] for _ in range(len(self.condition_names))
		]
		num_time_points = [
			0 for _ in range(len(self.condition_names))
		]
		for i in range(len(self.condition_names)):
			harvested_times[i] = list(sorted(set(harvested_times_reps[i])))
			num_time_points[i] = len(harvested_times[i])

		return harvested_times, harvested_times_reps, num_time_points

	def get_cell_number(self, worksheet: Worksheet) -> PerCond[PerTime[PerRep[PerGen[CellCount]]]]:
		"""
		Collects main data of the experiment.
		Sorts cell number data according to icnd, itpt, igen.

		:param1 worksheet: (Worksheet) openpyxl worksheet object
		:param2 data_format: (bool) validation of data format
		:returns: (list) cell number information
		:raises ArithmeticError: it's possible to encounter divide by 0 case
		"""

		# Define a mirrored list of harvest_time_reps to counts for number of replicates per condition per time point
		NUMBER_OF_REPLICATES = []
		for icnd in range(len(self.condition_names)):
			NUMBER_OF_REPLICATES.append(
				[len(list(group)) for _key, group in groupby(self.harvested_times_reps[icnd])])

		dataset = [
			[
				[
					[] for _ in range(max(self.generation_per_condition) + 1)
				] for _ in range(max(self.num_time_points))
			] for _ in range(len(self.condition_names))
		]

		icnd, itpt, igen = 0, 0, 0
		replicate_counter = 0
		time_row = 2
		next_condition_break = self.generation_per_condition[icnd] + 4
		for row_cell_object in worksheet.iter_rows(min_col=2, max_col=worksheet.max_column, min_row=2, max_row=worksheet.max_row):
			for cell in row_cell_object:
				val = cell.value
				if cell.row != time_row and cell.row != next_condition_break:
					if val is not None:
						dataset[icnd][itpt][igen].append(val)
						replicate_counter += 1
						if cell.col_idx == (len(self.harvested_times_reps[icnd]) + 1):
							igen += 1
						if replicate_counter % NUMBER_OF_REPLICATES[icnd][itpt] == 0:
							itpt += 1
							replicate_counter = 0
				if cell.col_idx == (len(self.harvested_times_reps[icnd]) + 1):
					itpt = 0
					if cell.row % next_condition_break == 0:
						time_row += self.generation_per_condition[icnd] + 3
						icnd += 1
						igen = 0
						if icnd > len(self.condition_names) - 1:
							break
						next_condition_break += self.generation_per_condition[icnd] + 3

		return remove_empty(dataset)


class ModernReadData(ReadData):
	meta_information: ExperimentInfo 
	
	@staticmethod
	def get_experiment_info(worksheet: Worksheet) -> ExperimentInfo:
		"""
		Collects user defined experiment information.

		:param1 worksheet: (Worksheet) openpyxl worksheet object
		:param2 data_format: (bool) True if we're using the new Excel data format
		:returns: (dict) experiments basic information
		:raises CustomError: raises a CustomError (inherits Exception) for missing critical information
		"""
		# Initialise information dictionary
		exp_info: ExperimentInfo = {
			'num_tp': 0,
			'last_gen': 0,
			'num_condition': 0,
			'num_replicate': 0,
			'num_beads': 0,
			'prop_bead_gated': 0
		}
		col_idx = 0

		# Iterate through rows [B3 - B8] where experiment informations are located
		# col_idx is essentially a check index
		for col_cell_object in worksheet.iter_rows(min_col=2, max_col=2, min_row=3, max_row=8):
			for cell in col_cell_object:
				header = cell.value
				# If cells are empty raise an Exception. This is to guide user to fix their data.
				if header is None:
					if col_idx == 0:
						raise Exception("Number of time points not set")
					elif col_idx == 1:
						raise Exception("Last generation not set")
					elif col_idx == 2:
						raise Exception("Number of conditions not set")
					elif col_idx == 3:
						raise Exception("Number of replicates not set")
					elif col_idx == 4:
						raise Exception("Number of beads not set")
					elif col_idx == 5:
						raise Exception("Prop. beads gated not set")
				else:
					# Validate if data make sense
					if col_idx == 0:
						if not isinstance(header, int):
							raise Exception("Number of time points must be an integer")
						if header <= 0:
							raise Exception("Number of time points cannot be 0 or negative")
						else:
							exp_info['num_tp'] = header
					elif col_idx == 1:
						if not isinstance(header, int):
							raise Exception("Last generation must be an integer")
						if header <= 0:
							raise Exception("Last generation must be greater than 0")
						else:
							exp_info['last_gen'] = header
					elif col_idx == 2:
						if not isinstance(header, int):
							raise Exception("Number of conditions must be an integer")
						if header <= 0:
							raise Exception("Number of condition must to be greater than 0")
						else:
							exp_info['num_condition'] = header
					elif col_idx == 3:
						if not isinstance(header, int):
							raise Exception("Number of replicates must be an integer")
						if header <= 0:
							raise Exception("Number of replicate must be greater than 0")
						else:
							exp_info['num_replicate'] = header
					elif col_idx == 4:
						if not isinstance(header, int):
							raise Exception("Number of beads must be an integer")
						if header <= 0:
							raise Exception("Number of beads must be greater than 0")
						else:
							exp_info['num_beads'] = header
					elif col_idx == 5:
						if not isinstance(header, float):
							raise Exception("Proportion of gated beads must be a float")
						if 0.0 <= header <= 1.0:
							exp_info['prop_bead_gated'] = header
						else:
							raise Exception("Prop. beads gated must be in range 0 and 1")
				col_idx += 1

		return exp_info

	@staticmethod
	def get_condition_names(worksheet: Worksheet) -> Conditions:
		"""
		Collects experiment condition names (e.g. Wildtype-1U, Wildtype-Unstim)

		:param1 worksheet: openpyxl worksheet object
		:param2 data_format: validation of data format
		:returns: condition names
		"""

		conditions = []
		cells = worksheet['A12':'A' + str(worksheet.max_row)]
		for cell in cells:
			val = cell.value
			if val is not None:
				# Converts to string in case of rare numeric condition name
				conditions.append(str(val))
			else:
				break
		return conditions

	def get_generation_information(self, worksheet: Worksheet) -> PerCond[MaxGeneration]:
		"""
		Collects last generation information per conditions.

		:param1 worksheet: (Worksheet) openpyxl worksheet object
		:param2 data_format: (bool) validation of data format
		:returns: (list) last generation per condition indexed by exact order of condition list
		"""

		generation = []
		# There is no way to identify last generation per condition in new data format.
		# Assume that all conditions are sharing same last generation that user set.
		for _idx in range(len(self.condition_names)):
			generation.append(self.meta_information['last_gen'])
		return generation

	def get_time_points(self, worksheet: Worksheet) -> tuple[
		PerCond[PerTime[HarvestTime]], PerCond[Reps[HarvestTime]], NumTimePoints
	]:
		"""
		Collects harvested time points per conditions.

		:param1 worksheet: openpyxl worksheet object
		:param2 data_format: validation of data format
		:returns: (tuple; list, list) returns harvested time points & duplicates of harvested time points to represent replicates
		"""

		# Initialise with empty list with same dimension as get_condition_names. This allows to access time point data sorted by condition index.
		harvested_times_reps = [
			[] for _ in range(len(self.condition_names))
		]

		cells = worksheet['D2':'D' + str(worksheet.max_row)]
		# Brings sample_names column from the data to check
		condition_column = worksheet['E2':'E' + str(worksheet.max_row)]
		for i, cell in enumerate(cells):
			val = cell.value
			curr_condition_name = str(condition_column[i].value)
			# If cell has a value means time to move onto next time point
			tmp = None
			if val is not None:
				tmp = val
				if curr_condition_name in self.condition_names:
					target_condition_idx = self.condition_names.index(curr_condition_name)
					harvested_times_reps[target_condition_idx].append(val)
			elif tmp is not None:
				if curr_condition_name in self.condition_names:
					target_condition_idx = self.condition_names.index(curr_condition_name)
					harvested_times_reps[target_condition_idx].append(tmp)
		harvested_times = [
			[] for _ in range(len(self.condition_names))
		]
		num_time_points = [
			0 for _ in range(len(self.condition_names))
		]
		# Remove duplicates in harvested_time_reps
		for i in range(len(self.condition_names)):
			harvested_times[i] = list(sorted(set(harvested_times_reps[i])))
			num_time_points[i] = len(harvested_times[i])

		return harvested_times, harvested_times_reps, num_time_points

	def get_cell_number(self, worksheet: Worksheet) -> PerCond[PerTime[PerRep[PerGen[CellCount]]]]:
		"""
		Collects main data of the experiment.
		Sorts cell number data according to icnd, itpt, igen.

		:param1 worksheet: (Worksheet) openpyxl worksheet object
		:param2 data_format: (bool) validation of data format
		:returns: (list) cell number information
		:raises ArithmeticError: it's possible to encounter divide by 0 case
		"""

		dataset = [
			[
				[
					[] for _ in range(self.meta_information['last_gen'] + 1)
				] for _ in range(self.meta_information['num_tp'])
			] for _ in range(len(self.condition_names))
		]

		exp_beads = self.meta_information['num_beads'] * self.meta_information['prop_bead_gated']
		itpt = 0
		row_counter = 0
		time_list = worksheet['D2':'D' + str(worksheet.max_row)]
		condition_list = worksheet['E2':'E' + str(worksheet.max_row)]
		bead_list = worksheet['F2':'F' + str(worksheet.max_row)]
		for row_cell_object in worksheet.iter_rows(min_col=8, max_col=worksheet.max_column, min_row=2, max_row=worksheet.max_row):
			for col_idx, cell in enumerate(row_cell_object):
				val = cell.value
				curr_condition_name = str(condition_list[row_counter].value)
				curr_time = time_list[row_counter].value
				curr_beads = bead_list[row_counter].value
				# Time point update criteria
				if curr_time is not None and row_counter > 0 and col_idx == 0:
					itpt += 1
				# Iteratively check if condition name exists in self.condition_name.
				# This way user can manually take out one of replicates without deleteing entire row, and system will dynamically adjust matrix size
				if curr_condition_name in self.condition_names:
					target_condition_idx = self.condition_names.index(curr_condition_name)
					if curr_beads is None:
						raise ArithmeticError(
							"Empty bead number is detected. Check your number of beads at : time point %.2f, condition %s" % (
								self.harvested_times[target_condition_idx][itpt], curr_condition_name))
					if isinstance(val, SupportsFloat) and isinstance(curr_beads, SupportsFloat):
						if float(curr_beads) == 0:
							raise ArithmeticError(
								"Divide by zero issue detected. Check your number of beads at : time point %.2f, condition %s" % (
									self.harvested_times[target_condition_idx][itpt], curr_condition_name))
						else:
							val = float(val) * exp_beads / float(curr_beads)
					else:
						val = 0
					dataset[target_condition_idx][itpt][col_idx].append(val)

				if col_idx == self.meta_information['last_gen']:
					row_counter += 1

		return remove_empty(dataset)
